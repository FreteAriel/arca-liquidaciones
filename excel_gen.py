"""
Generador de Excel: Libro IVA Ventas y Libro IVA Compras
Cruza datos de ARCA (comprobantes + retenciones SICORE) con ARBA (deducciones)
"""

import os
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# -----------------------------------------------------------------------
# Colores y estilos
# -----------------------------------------------------------------------

COLOR_HEADER_VENTAS   = "1F4E79"  # azul oscuro
COLOR_HEADER_COMPRAS  = "375623"  # verde oscuro
COLOR_HEADER_TEXT     = "FFFFFF"
COLOR_FILA_PAR        = "EBF3FB"
COLOR_TOTAL_ROW       = "FFF2CC"

def _border_thin():
    lado = Side(style="thin", color="AAAAAA")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _header_style(ws, row, col, value, color_fondo):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=True, color=COLOR_HEADER_TEXT, size=10)
    cell.fill = PatternFill("solid", fgColor=color_fondo)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border_thin()
    return cell

def _data_style(ws, row, col, value, fmt=None, bold=False, bg=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=9, bold=bold)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border_thin()
    if fmt:
        cell.number_format = fmt
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


# -----------------------------------------------------------------------
# Notas de Crédito — deben restar en el libro
# -----------------------------------------------------------------------

def _es_nota_credito(tipo: str) -> bool:
    """
    True si el comprobante es una Nota de Crédito.
    Aplica tanto a emitidos (ventas) como a recibidos (compras).
    Los tipos ARCA incluyen p.ej.: '3 - Nota de Crédito B',
    '8 - Nota de Crédito A', '13 - Nota de Crédito C', etc.
    """
    t = (tipo or "").lower()
    # normalizar tildes
    t = t.replace("é", "e").replace("é", "e").replace("ó", "o")
    return "nota de cr" in t


def _signo(tipo: str) -> float:
    """Devuelve -1.0 para Notas de Crédito, +1.0 para el resto."""
    return -1.0 if _es_nota_credito(tipo) else 1.0


CAMPOS_MONETARIOS_VENTAS  = ["neto_21","iva_21","neto_105","iva_105","exento","total"]
CAMPOS_MONETARIOS_COMPRAS = [
    "neto_21","iva_21","neto_105","iva_105","neto_27","iva_27",
    "ret_iibb","ret_ganancias","ret_suss","ret_iva","exento","total",
]


# -----------------------------------------------------------------------
# Exportar LIBRO IVA VENTAS
# -----------------------------------------------------------------------

COLS_VENTAS = [
    ("FECHA",              12),
    ("TIPO COMPROBANTE",   20),
    ("N° COMPROBANTE",     18),
    ("CUIT",               16),
    ("RAZÓN SOCIAL",       30),
    ("NETO 21%",           14),
    ("IVA 21%",            14),
    ("NETO 10,5%",         14),
    ("IVA 10,5%",          14),
    ("EXENTO",             14),
    ("TOTAL",              14),
]

def _escribir_ventas(ws, comprobantes: list[dict]):
    color = COLOR_HEADER_VENTAS
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 35

    for col_idx, (titulo, ancho) in enumerate(COLS_VENTAS, 1):
        _header_style(ws, 1, col_idx, titulo, color)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    totales = {k: 0.0 for k in ["neto_21","iva_21","neto_105","iva_105","exento","total"]}
    fmt_num = '#,##0.00'
    fmt_fecha = '@'

    for i, comp in enumerate(comprobantes, 2):
        bg = COLOR_FILA_PAR if i % 2 == 0 else None
        tipo_comp = comp.get("tipo","")
        sgn = _signo(tipo_comp)
        _data_style(ws, i, 1,  comp.get("fecha",""),            bg=bg)
        _data_style(ws, i, 2,  tipo_comp,                       bg=bg)
        _data_style(ws, i, 3,  comp.get("numero",""),           bg=bg)
        _data_style(ws, i, 4,  comp.get("cuit_contraparte",""), bg=bg)
        _data_style(ws, i, 5,  comp.get("razon_social",""),     bg=bg)
        for j, key in enumerate(["neto_21","iva_21","neto_105","iva_105","exento","total"], 6):
            if key == "total":
                # Fórmula Excel: suma de netos+IVAs+exento (F:J), ya con signo aplicado
                formula = f"=SUM(F{i}:J{i})"
                _data_style(ws, i, j, formula, fmt=fmt_num, bg=bg)
                totales[key] += sgn * (comp.get("neto_21",0.0) + comp.get("iva_21",0.0) +
                                       comp.get("neto_105",0.0) + comp.get("iva_105",0.0) +
                                       comp.get("exento",0.0))
            else:
                val = sgn * (comp.get(key, 0.0) or 0.0)
                _data_style(ws, i, j, val, fmt=fmt_num, bg=bg)
                totales[key] += val

    # Fila de totales verticales — fórmulas =SUM(col2:col_ultima_fila_datos)
    fila_total = len(comprobantes) + 2
    primera_fila_datos = 2
    _data_style(ws, fila_total, 5, "TOTAL", bold=True, bg=COLOR_TOTAL_ROW)
    for j, key in enumerate(["neto_21","iva_21","neto_105","iva_105","exento","total"], 6):
        col_letra = get_column_letter(j)
        formula_vert = f"=SUM({col_letra}{primera_fila_datos}:{col_letra}{fila_total - 1})"
        _data_style(ws, fila_total, j, formula_vert, fmt=fmt_num, bold=True, bg=COLOR_TOTAL_ROW)

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS_VENTAS))}{len(comprobantes)+1}"


# -----------------------------------------------------------------------
# Exportar LIBRO IVA COMPRAS
# -----------------------------------------------------------------------

COLS_COMPRAS = [
    ("FECHA",                  12),
    ("TIPO COMPROBANTE",       20),
    ("N° COMPROBANTE",         18),
    ("CUIT",                   16),
    ("RAZÓN SOCIAL",           30),
    ("NETO 21%",               14),
    ("IVA 21%",                14),
    ("NETO 10,5%",             14),
    ("IVA 10,5%",              14),
    ("NETO 27%",               14),
    ("IVA 27%",                14),
    ("RET. ING. BRUTOS",       16),
    ("RET. GANANCIAS",         16),
    ("RET. SUSS",              14),
    ("RET. IVA",               14),
    ("EXENTO",                 14),
    ("TOTAL",                  14),
]

def _escribir_compras(ws, comprobantes: list[dict]):
    color = COLOR_HEADER_COMPRAS
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 35

    for col_idx, (titulo, ancho) in enumerate(COLS_COMPRAS, 1):
        _header_style(ws, 1, col_idx, titulo, color)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    COLS_NUM = [
        # (clave_dict, columna_excel)
        ("neto_21", 6), ("iva_21", 7), ("neto_105", 8), ("iva_105", 9),
        ("neto_27", 10), ("iva_27", 11),
        ("ret_iibb", 12), ("ret_ganancias", 13), ("ret_suss", 14),
        ("ret_iva", 15), ("exento", 16), ("total", 17),
    ]
    # TOTAL incluye TODAS las columnas: netos + IVAs + retenciones + exento (F a P)
    KEYS_PARA_TOTAL = {"neto_21","iva_21","neto_105","iva_105","neto_27","iva_27",
                       "ret_iibb","ret_ganancias","ret_suss","ret_iva","exento"}

    totales = {k: 0.0 for k, _ in COLS_NUM}
    fmt_num = '#,##0.00'

    for i, comp in enumerate(comprobantes, 2):
        bg = COLOR_FILA_PAR if i % 2 == 0 else None
        tipo_comp = comp.get("tipo", "")
        sgn = _signo(tipo_comp)
        _data_style(ws, i, 1,  comp.get("fecha",""),            bg=bg)
        _data_style(ws, i, 2,  tipo_comp,                       bg=bg)
        _data_style(ws, i, 3,  comp.get("numero",""),           bg=bg)
        _data_style(ws, i, 4,  comp.get("cuit_contraparte",""), bg=bg)
        _data_style(ws, i, 5,  comp.get("razon_social",""),     bg=bg)

        # Aplicar signo (−1 para Notas de Crédito) y escribir valores
        # TOTAL como fórmula Excel =SUM(F:K,P) → recalculable si el usuario edita celdas
        vals = {key: sgn * (comp.get(key, 0.0) or 0.0) for key, _ in COLS_NUM}

        for key, col in COLS_NUM:
            if key == "total":
                # Fórmula Excel: suma TODAS las columnas F a P (netos + IVAs + retenciones + exento)
                formula = f"=SUM(F{i}:P{i})"
                _data_style(ws, i, col, formula, fmt=fmt_num, bg=bg)
                totales[key] += sum(vals[k] for k in vals if k != "total")
            else:
                _data_style(ws, i, col, vals[key], fmt=fmt_num, bg=bg)
                totales[key] += vals[key]

    # Fila de totales verticales — fórmulas =SUM(col2:col_ultima_fila_datos)
    fila_total = len(comprobantes) + 2
    primera_fila_datos = 2
    _data_style(ws, fila_total, 5, "TOTAL", bold=True, bg=COLOR_TOTAL_ROW)
    for key, col in COLS_NUM:
        col_letra = get_column_letter(col)
        formula_vert = f"=SUM({col_letra}{primera_fila_datos}:{col_letra}{fila_total - 1})"
        _data_style(ws, fila_total, col, formula_vert, fmt=fmt_num, bold=True, bg=COLOR_TOTAL_ROW)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS_COMPRAS))}{len(comprobantes)+1}"


# -----------------------------------------------------------------------
# Cruzar retenciones SICORE y ARBA con comprobantes recibidos
# -----------------------------------------------------------------------

def _cruzar_retenciones(comprobantes_recibidos: list[dict],
                         retenciones_sicore: list[dict],
                         deducciones_arba: dict,
                         log_fn=None) -> list[dict]:
    """
    Para cada comprobante recibido, cruza retenciones SICORE 767 (IVA) y deducciones ARBA.

    Lógica SICORE 767:
      - Todas las retenciones de SICORE 767 son de IVA → van a ret_iva
      - Match principal: CUIT agente == CUIT contraparte del comprobante
      - Cada retencion individual se cruza 1-a-1 por CUIT (si un proveedor tiene
        múltiples retenciones, se suman)

    Lógica ARBA:
      - percepciones + sitrac + retenciones_bancarias → ret_iibb

    Si una retención SICORE no encuentra comprobante coincidente → se agrega
    como fila extra al final (con campos de monto en cero excepto ret_iva).
    """

    def _log(msg):
        if log_fn:
            log_fn(msg)

    def _extraer_num(nro: str) -> int:
        """
        Extrae el número de factura (NUM) de cualquier formato de comprobante.
        - IVA COMPRAS '00014-00038247'  → últimos 8 del part derecho → 38247
        - IVA COMPRAS '00029-00032701'  → 32701
        - SICORE 16d  '0001400000038247'→ últimos 8 dígitos → 38247
        - SICORE 16d  '0000002900032701'→ últimos 8 dígitos → 32701
        - SICORE 13d  '0701100115645'   → últimos 8 dígitos → 115645? skip match
        Los últimos 8 dígitos del string de dígitos siempre son el NUM.
        """
        digits = re.sub(r'[^0-9]', '', str(nro).strip())
        if '-' in str(nro):
            # Formato IVA COMPRAS: XXXXX-YYYYYYYY → tomar parte derecha
            try:
                return int(str(nro).split('-', 1)[1])
            except Exception:
                pass
        if len(digits) >= 8:
            return int(digits[-8:])
        return -1

    def _clave_match(cuit: str, nro: str) -> tuple:
        """Clave: (CUIT_limpio, NUM_factura) — coincide entre formatos IVA y SICORE."""
        return (re.sub(r'[^0-9]', '', str(cuit)), _extraer_num(nro))

    def _normalizar_rs(rs: str) -> str:
        """Normaliza razón social para comparación aproximada."""
        import unicodedata
        rs = str(rs).upper().strip()
        for sufijo in ['S.A.', 'S.R.L.', 'S.A.S.', 'SRL', 'SAU', 'SAS', 'SA']:
            rs = rs.replace(sufijo, '')
        rs = ''.join(
            c for c in unicodedata.normalize('NFD', rs)
            if unicodedata.category(c) != 'Mn'
        )
        return re.sub(r'\s+', ' ', rs).strip()

    # ── Indexar SICORE por (CUIT_agente, NUM) y por NUM solo ─────────────
    sicore_por_clave: dict[tuple, list[dict]] = {}
    sicore_por_num: dict[int, list[dict]] = {}
    for ret in retenciones_sicore:
        clave = _clave_match(ret.get("cuit_agente", ""), ret.get("numero_comprobante", ""))
        sicore_por_clave.setdefault(clave, []).append(ret)
        num = _extraer_num(ret.get("numero_comprobante", ""))
        if num > 0:
            sicore_por_num.setdefault(num, []).append(ret)

    _log(f"   🔗 SICORE: {len(retenciones_sicore)} retenciones indexadas por (CUIT, NUM) y por NUM")

    # ── Indexar solo PERCEPCIONES CP por (CUIT, NUM) y por NUM-solo ──────
    # CP: tienen numero_formateado "PPPPP-NNNNNNNN" → match per-factura preciso.
    # CT (SITRAC) y CB (Bancarias): sin número de factura → van como filas extra al final.
    arba_por_clave: dict[tuple, float] = {}   # (cuit_agente, num) → importe CP
    arba_por_cuit:  dict[str,   float] = {}   # cuit_agente → importe CP total (fallback)

    for item in deducciones_arba.get("percepciones", []):
        cuit   = re.sub(r'[^0-9]', '', item.get("cuit_agente", ""))
        num    = _extraer_num(item.get("numero_formateado", "") or str(item.get("numero", "")))
        importe = item.get("importe", 0.0)
        if cuit and num > 0:
            clave = (cuit, num)
            arba_por_clave[clave] = arba_por_clave.get(clave, 0.0) + importe
        if cuit:
            arba_por_cuit[cuit] = arba_por_cuit.get(cuit, 0.0) + importe

    _log(f"   🔗 ARBA CP: {len(arba_por_clave)} claves (CUIT+NUM), {len(arba_por_cuit)} CUITs")

    # Índice por NUM solo (fallback nivel 3): solo si el NUM es único en ARBA
    from collections import Counter
    _num_count = Counter(num for (_, num) in arba_por_clave.keys())
    arba_por_num: dict[int, float] = {
        num: imp
        for (cuit, num), imp in arba_por_clave.items()
        if _num_count[num] == 1  # único → match seguro sin CUIT
    }
    _log(f"   🔗 ARBA CP NUM-solo: {len(arba_por_num)} claves únicas")

    # ── Cruzar comprobantes recibidos con SICORE ─────────────────────────
    matched_ret_ids: set[int] = set()   # id() de cada ret dict ya machado
    matched_arba_keys: set[tuple] = set()  # (cuit, num) CP ya matcheados
    resultado: list[dict] = []
    arba_matches = 0  # contador de matches ARBA exitosos

    for comp in comprobantes_recibidos:
        cuit_prov = re.sub(r'[^0-9]', '', comp.get("cuit_contraparte", ""))
        clave     = _clave_match(cuit_prov, comp.get("numero", ""))
        comp_out  = dict(comp)

        # ARBA → ret_iibb:
        #   Nivel 1: match exacto (CUIT_proveedor + NUM_factura) ← más preciso
        #   Nivel 2: fallback CUIT-solo (cubre SITRAC / bancarias sin num)
        #   Nivel 3: fallback NUM-solo (cubre el caso donde CUIT no llegó bien de ARCA)
        num_comp = clave[1]
        ret_iibb = arba_por_clave.get((cuit_prov, num_comp), None)
        if ret_iibb is not None:
            _log(f"   ✅ ARBA L1 CUIT+NUM: {cuit_prov} N°{num_comp} → {ret_iibb:.2f}")
            matched_arba_keys.add((cuit_prov, num_comp))
            arba_matches += 1
        elif num_comp > 0 and num_comp in arba_por_num:
            ret_iibb = arba_por_num[num_comp]
            _log(f"   ℹ️ ARBA L3 NUM-solo: N°{num_comp} → {ret_iibb:.2f}")
            # Marcar la clave real como matcheada
            for (c, n) in arba_por_clave:
                if n == num_comp:
                    matched_arba_keys.add((c, n))
                    break
            arba_matches += 1
        else:
            ret_iibb = 0.0
        comp_out["ret_iibb"] = ret_iibb
        comp_out.setdefault("ret_ganancias", 0.0)
        comp_out.setdefault("ret_suss", 0.0)

        # ── Nivel 1: match exacto por (CUIT, NUM) ──────────────────────
        rets_match = sicore_por_clave.get(clave, [])
        match_tipo = "CUIT+NUM"

        # ── Nivel 2 fallback: match por NUM + razón social ─────────────
        #    Cubre casos donde el CUIT del comprobante (ej. persona 20-...)
        #    difiere del CUIT agente SICORE (ej. empresa 30-...) pero es
        #    la misma entidad.
        if not rets_match:
            num_comp = clave[1]
            candidatos = sicore_por_num.get(num_comp, [])
            if candidatos:
                rs_comp = _normalizar_rs(comp.get("razon_social", ""))
                for candidato in candidatos:
                    rs_ret = _normalizar_rs(candidato.get("razon_social_agente", ""))
                    if rs_comp and rs_ret and (
                        rs_comp in rs_ret or rs_ret in rs_comp or
                        rs_comp[:8] == rs_ret[:8]
                    ):
                        rets_match.append(candidato)
                        match_tipo = "NUM+RS"

        ret_iva = sum(r.get("importe", 0.0) for r in rets_match)
        comp_out["ret_iva"] = ret_iva

        if rets_match:
            for r in rets_match:
                matched_ret_ids.add(id(r))
            rs_comp   = _normalizar_rs(comp.get("razon_social", ""))
            rs_sicore = _normalizar_rs(rets_match[0].get("razon_social_agente", ""))
            rs_ok = rs_comp and rs_sicore and (
                rs_comp in rs_sicore or rs_sicore in rs_comp or
                rs_comp[:10] == rs_sicore[:10]
            )
            icono = "✅" if rs_ok else "⚠️ RS difiere"
            _log(f"   {icono} Match ({match_tipo}) NUM {clave[1]} → "
                 f"ret_iva={ret_iva:.2f} | '{comp.get('razon_social','')}' ↔ '{rets_match[0].get('razon_social_agente','')}'")

        resultado.append(comp_out)

    _log(f"   📊 ARBA CP: {arba_matches} matches sobre {len(comprobantes_recibidos)} comprobantes")

    # ── CP sin match → filas extra (percepciones sin comprobante en ARCA) ─
    percepciones_items = deducciones_arba.get("percepciones", [])
    # Reconstruir mapa clave→item para encontrar los no matcheados
    _cp_items: dict[tuple, dict] = {}
    for item in percepciones_items:
        cuit_a = re.sub(r'[^0-9]', '', item.get("cuit_agente", ""))
        num_a  = _extraer_num(item.get("numero_formateado", "") or str(item.get("numero", "")))
        if cuit_a and num_a > 0:
            _cp_items[(cuit_a, num_a)] = item

    cp_sin_match = [(k, v) for k, v in arba_por_clave.items() if k not in matched_arba_keys]
    if cp_sin_match:
        _log(f"   📋 ARBA CP sin match: {len(cp_sin_match)} percepciones sin factura ARCA → filas extra")
        for (cuit, num), imp in cp_sin_match:
            item_orig = _cp_items.get((cuit, num), {})
            nro_fmt   = item_orig.get("numero_formateado", f"{num:08d}")
            resultado.append({
                "fecha":            item_orig.get("fecha", ""),
                "tipo":             "PERCEPCION IIBB",
                "numero":           nro_fmt,
                "cuit_contraparte": cuit,
                "razon_social":     cuit,
                "neto_21": 0.0, "iva_21": 0.0,
                "neto_105": 0.0, "iva_105": 0.0,
                "neto_27": 0.0, "iva_27": 0.0,
                "ret_iibb": imp,
                "ret_ganancias": 0.0, "ret_suss": 0.0, "ret_iva": 0.0,
                "exento": 0.0, "total": 0.0,
            })

    # ── CP con num=0 (sin número de comprobante válido) → también filas extra
    cp_num_cero = [item for item in percepciones_items
                   if _extraer_num(item.get("numero_formateado", "") or str(item.get("numero", ""))) == 0
                   and re.sub(r'[^0-9]', '', item.get("cuit_agente", ""))]
    if cp_num_cero:
        _log(f"   📋 ARBA CP num=0: {len(cp_num_cero)} percepciones sin número → filas extra")
        for item in cp_num_cero:
            cuit = re.sub(r'[^0-9]', '', item.get("cuit_agente", ""))
            resultado.append({
                "fecha":            item.get("fecha", ""),
                "tipo":             "PERCEPCION IIBB",
                "numero":           item.get("numero_formateado", ""),
                "cuit_contraparte": cuit,
                "razon_social":     cuit,
                "neto_21": 0.0, "iva_21": 0.0,
                "neto_105": 0.0, "iva_105": 0.0,
                "neto_27": 0.0, "iva_27": 0.0,
                "ret_iibb": item.get("importe", 0.0),
                "ret_ganancias": 0.0, "ret_suss": 0.0, "ret_iva": 0.0,
                "exento": 0.0, "total": 0.0,
            })

    # ── SITRAC (CT) → filas extra al final ───────────────────────────────
    sitrac_items = deducciones_arba.get("sitrac", [])
    if sitrac_items:
        _log(f"   📋 ARBA SITRAC: {len(sitrac_items)} registros → filas extra")
        for item in sitrac_items:
            cuit = re.sub(r'[^0-9]', '', item.get("cuit_agente", ""))
            resultado.append({
                "fecha":            item.get("fecha", ""),
                "tipo":             "SIRTAC IIBB",
                "numero":           "",
                "cuit_contraparte": cuit,
                "razon_social":     item.get("cuit_agente", ""),
                "neto_21": 0.0, "iva_21": 0.0,
                "neto_105": 0.0, "iva_105": 0.0,
                "neto_27": 0.0, "iva_27": 0.0,
                "ret_iibb": item.get("importe", 0.0),
                "ret_ganancias": 0.0, "ret_suss": 0.0, "ret_iva": 0.0,
                "exento": 0.0, "total": 0.0,
            })

    # ── Bancarias (CB) → filas extra al final ────────────────────────────
    bancarias_items = deducciones_arba.get("retenciones_bancarias", [])
    if bancarias_items:
        _log(f"   📋 ARBA Bancarias: {len(bancarias_items)} registros → filas extra")
        for item in bancarias_items:
            cuit = re.sub(r'[^0-9]', '', item.get("cuit_agente", ""))
            resultado.append({
                "fecha":            item.get("fecha", ""),
                "tipo":             "RET. BANCARIA IIBB",
                "numero":           item.get("tipo_doc", ""),
                "cuit_contraparte": cuit,
                "razon_social":     item.get("cuit_agente", ""),
                "neto_21": 0.0, "iva_21": 0.0,
                "neto_105": 0.0, "iva_105": 0.0,
                "neto_27": 0.0, "iva_27": 0.0,
                "ret_iibb": item.get("importe", 0.0),
                "ret_ganancias": 0.0, "ret_suss": 0.0, "ret_iva": 0.0,
                "exento": 0.0, "total": 0.0,
            })

    # ── Retenciones SICORE sin comprobante coincidente → filas extra ─────
    sin_match_rets = [r for r in retenciones_sicore if id(r) not in matched_ret_ids]
    _log(f"   ⚠️ SICORE sin match: {len(sin_match_rets)} retenciones → se agregan al final")

    for ret in sin_match_rets:
        resultado.append({
            "fecha":            ret.get("fecha_comprobante") or ret.get("fecha", ""),
            "tipo":             "RETENCIÓN SICORE 767",
            "numero":           ret.get("numero_comprobante", ""),
            "cuit_contraparte": re.sub(r'[^0-9]', '', ret.get("cuit_agente", "")),
            "razon_social":     ret.get("razon_social_agente", ""),
            "neto_21": 0.0, "iva_21": 0.0,
            "neto_105": 0.0, "iva_105": 0.0,
            "neto_27": 0.0, "iva_27": 0.0,
            "ret_iibb": 0.0, "ret_ganancias": 0.0, "ret_suss": 0.0,
            "ret_iva": ret.get("importe", 0.0),
            "exento": 0.0, "total": 0.0,
        })

    return resultado


# -----------------------------------------------------------------------
# Función principal
# -----------------------------------------------------------------------

def generar_excel(
    comprobantes_emitidos: list[dict],
    comprobantes_recibidos: list[dict],
    retenciones_sicore: list[dict],
    deducciones_arba: dict,
    ruta_salida: str,
    saldo_favor_1p: float = 0.0,
    saldo_favor_2p: float = 0.0,
    alicuota_iibb: float = 0.0,
    saldo_anterior_iibb: float = 0.0,
    log_fn=None,
) -> str:
    """
    Genera el archivo Excel con cuatro hojas:
    - IVA VENTAS
    - IVA COMPRAS
    - POSICION DE IVA
    - POSICION DE INGRESOS BRUTOS
    """
    wb = openpyxl.Workbook()

    # --- Hoja IVA Ventas ---
    ws_ventas = wb.active
    ws_ventas.title = "IVA VENTAS"
    _escribir_ventas(ws_ventas, comprobantes_emitidos)

    # --- Hoja IVA Compras ---
    ws_compras = wb.create_sheet("IVA COMPRAS")
    compras_con_ret = _cruzar_retenciones(
        comprobantes_recibidos, retenciones_sicore, deducciones_arba, log_fn=log_fn
    )
    _escribir_compras(ws_compras, compras_con_ret)

    # ── Verificación RET. ING. BRUTOS vs totales ARBA ────────────────────
    _verificar_ret_iibb(compras_con_ret, deducciones_arba, log_fn=log_fn)

    # ── Calcular totales para hojas de posición ───────────────────────────
    def _sumar(lista, *campos):
        return sum(r.get(c, 0.0) or 0.0 for r in lista for c in campos if c in r)

    # Ventas: aplicar signo por nota de crédito
    iva21_ventas  = sum(_signo(r.get("tipo","")) * (r.get("iva_21",0) or 0) for r in comprobantes_emitidos)
    iva105_ventas = sum(_signo(r.get("tipo","")) * (r.get("iva_105",0) or 0) for r in comprobantes_emitidos)
    neto21_ventas  = sum(_signo(r.get("tipo","")) * (r.get("neto_21",0) or 0) for r in comprobantes_emitidos)
    neto105_ventas = sum(_signo(r.get("tipo","")) * (r.get("neto_105",0) or 0) for r in comprobantes_emitidos)

    # Compras: aplicar signo por nota de crédito (igual que ventas)
    # _cruzar_retenciones NO pre-niega; el signo lo aplica _escribir_compras al volcar al Excel
    iva21_compras   = sum(_signo(r.get("tipo","")) * (r.get("iva_21",0) or 0) for r in compras_con_ret)
    iva105_compras  = sum(_signo(r.get("tipo","")) * (r.get("iva_105",0) or 0) for r in compras_con_ret)
    ret_iva_compras  = sum((r.get("ret_iva",0) or 0) for r in compras_con_ret)   # retenciones siempre positivas
    ret_iibb_compras = sum((r.get("ret_iibb",0) or 0) for r in compras_con_ret)  # retenciones siempre positivas

    # --- Hoja POSICION DE IVA ---
    ws_iva_pos = wb.create_sheet("POSICION DE IVA")
    _escribir_posicion_iva(
        ws_iva_pos,
        debito_fiscal   = iva21_ventas + iva105_ventas,
        credito_fiscal  = iva21_compras + iva105_compras,
        saldo_favor_1p  = saldo_favor_1p,
        saldo_favor_2p  = saldo_favor_2p,
        ret_iva_total   = ret_iva_compras,
    )

    # --- Hoja POSICION DE INGRESOS BRUTOS ---
    ws_iibb_pos = wb.create_sheet("POSICION DE INGRESOS BRUTOS")
    _escribir_posicion_iibb(
        ws_iibb_pos,
        ingresos_declarados  = neto21_ventas + neto105_ventas,
        alicuota_pct         = alicuota_iibb,
        total_deducciones    = ret_iibb_compras,
        saldo_anterior_iibb  = saldo_anterior_iibb,
    )

    # Guardar
    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
    wb.save(ruta_salida)
    return ruta_salida


def _escribir_posicion_iva(
    ws,
    debito_fiscal: float,
    credito_fiscal: float,
    saldo_favor_1p: float,
    saldo_favor_2p: float,
    ret_iva_total: float,
):
    """
    Hoja POSICION DE IVA — replica la estructura del F.2051.

    Determinación del impuesto
    ──────────────────────────
    Total del débito fiscal del período          = IVA 21% + IVA 10,5% (Libro IVA Ventas)
    Total del crédito fiscal del período         = IVA 21% + IVA 10,5% (Libro IVA Compras)
    Saldo técnico a favor contribuyente anterior = saldo_favor_1p  (ingresado por usuario)
    Saldo técnico a favor del contribuyente      = abs(resultado) si resultado < 0

    Determinación de la posición mensual
    ─────────────────────────────────────
    Saldo técnico a favor de ARCA               = resultado si resultado > 0
    Saldo técnico a favor del contribuyente     = igual al ítem de la sección anterior
    Saldo libre disponibilidad anterior (2P)    = saldo_favor_2p  (ingresado por usuario)
    Total retenciones / percepciones / pagos    = RET. IVA del Libro IVA Compras
    Saldo libre disponibilidad del período      = 2P + retenciones
    """

    COLOR_HEADER  = "1F4E79"   # azul oscuro (= IVA Ventas)
    COLOR_SECCION = "D6E4F0"   # celeste suave
    COLOR_TOTAL   = "FFF2CC"   # amarillo suave
    COLOR_FAVOR   = "E2EFDA"   # verde suave  → saldo a favor del contribuyente
    COLOR_ARCA    = "FCE4D6"   # rojo suave   → saldo a favor de ARCA

    fmt_num = '#,##0.00'

    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 20

    fila = [1]   # lista mutable para poder modificar desde funciones internas

    def _titulo(texto):
        r = fila[0]
        ws.row_dimensions[r].height = 36
        cell = ws.cell(row=r, column=1, value=texto)
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border    = _border_thin()
        ws.merge_cells(f"A{r}:B{r}")
        fila[0] += 1

    def _subtitulo(texto):
        r = fila[0]
        ws.row_dimensions[r].height = 22
        cell = ws.cell(row=r, column=1, value=texto)
        cell.font      = Font(name="Calibri", bold=True, color="1F4E79", size=10)
        cell.fill      = PatternFill("solid", fgColor=COLOR_SECCION)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border    = _border_thin()
        ws.merge_cells(f"A{r}:B{r}")
        fila[0] += 1

    def _item(label, valor, bold=False, bg=None):
        r = fila[0]
        ws.row_dimensions[r].height = 20
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font      = Font(name="Calibri", size=10, bold=bold)
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        c1.border    = _border_thin()
        if bg:
            c1.fill = PatternFill("solid", fgColor=bg)

        c2 = ws.cell(row=r, column=2, value=valor)
        c2.font         = Font(name="Calibri", size=10, bold=bold)
        c2.number_format = fmt_num
        c2.alignment    = Alignment(horizontal="right", vertical="center")
        c2.border       = _border_thin()
        if bg:
            c2.fill = PatternFill("solid", fgColor=bg)
        fila[0] += 1

    def _espacio():
        fila[0] += 1

    # ── Cálculos ──────────────────────────────────────────────────────────
    resultado_tecnico = debito_fiscal - credito_fiscal - saldo_favor_1p
    if resultado_tecnico < 0:
        saldo_tec_contribuyente = abs(resultado_tecnico)
        saldo_tec_arca          = 0.0
    else:
        saldo_tec_contribuyente = 0.0
        saldo_tec_arca          = resultado_tecnico

    saldo_libre_disp = saldo_favor_2p + ret_iva_total

    # ── Construcción de la hoja ───────────────────────────────────────────
    _titulo("POSICIÓN DE IVA  ·  F.2051  ·  Determinación del Impuesto al Valor Agregado")

    _subtitulo("Determinación del impuesto")
    _item("Total del débito fiscal del período",                              debito_fiscal)
    _item("Total del crédito fiscal del período",                             credito_fiscal)
    _item("Saldo técnico a favor del contribuyente del período anterior",     saldo_favor_1p)
    _item(
        "Saldo técnico a favor del contribuyente",
        saldo_tec_contribuyente,
        bold=saldo_tec_contribuyente > 0,
        bg=COLOR_FAVOR if saldo_tec_contribuyente > 0 else None,
    )

    _espacio()

    _subtitulo("Determinación de la posición mensual")
    _item(
        "Saldo técnico a favor de ARCA",
        saldo_tec_arca,
        bold=saldo_tec_arca > 0,
        bg=COLOR_ARCA if saldo_tec_arca > 0 else None,
    )
    _item(
        "Saldo técnico a favor del contribuyente",
        saldo_tec_contribuyente,
        bold=saldo_tec_contribuyente > 0,
        bg=COLOR_FAVOR if saldo_tec_contribuyente > 0 else None,
    )
    _item("Saldo a favor de libre disponibilidad del período anterior neto de usos",
          saldo_favor_2p)
    _item("Total de retenciones, percepciones y pagos a cuenta neto de restituciones",
          ret_iva_total)
    _item(
        "Saldo de libre disponibilidad a favor del contribuyente del período",
        saldo_libre_disp,
        bold=True,
        bg=COLOR_TOTAL,
    )


def _escribir_posicion_iibb(
    ws,
    ingresos_declarados: float,
    alicuota_pct: float,
    total_deducciones: float,
    saldo_anterior_iibb: float,
):
    """
    Hoja POSICION DE INGRESOS BRUTOS — replica la estructura del R-606M de ARBA.

    Ingresos / impuesto
    ────────────────────
    Ingresos declarados          = Neto 21% + Neto 10,5% del Libro IVA Ventas
    Alícuota IIBB (%)            = ingresada por el usuario
    Impuesto determinado         = Ingresos × Alícuota / 100

    Deducciones
    ────────────
    Total deducciones            = RET. ING. BRUTOS (Libro IVA Compras)

    Posición del período
    ─────────────────────
    Saldo de la DJ               = Impuesto determinado − Total deducciones
    Saldo a favor anterior IIBB  = ingresado por el usuario (valor positivo)
    Saldo acumulado al cierre    = Saldo de la DJ − Saldo a favor anterior
                                   (el saldo a favor anterior reduce la deuda)

    Resultado
    ──────────
    Si saldo_acumulado > 0  → A PAGAR IIBB  (el impuesto supera los créditos)
    Si saldo_acumulado ≤ 0  → IIBB A FAVOR  (los créditos superan el impuesto)
    """

    COLOR_HEADER  = "375623"   # verde oscuro (= IVA Compras)
    COLOR_SECCION = "E2EFDA"   # verde suave
    COLOR_TOTAL   = "FFF2CC"   # amarillo suave
    COLOR_PAGAR   = "FCE4D6"   # rojo suave  → A PAGAR
    COLOR_FAVOR   = "E2EFDA"   # verde suave → A FAVOR

    fmt_num = '#,##0.00'
    fmt_pct = '0.00"%"'        # muestra el número tal cual con símbolo %

    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 20

    fila = [1]

    def _titulo(texto):
        r = fila[0]
        ws.row_dimensions[r].height = 36
        cell = ws.cell(row=r, column=1, value=texto)
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border    = _border_thin()
        ws.merge_cells(f"A{r}:B{r}")
        fila[0] += 1

    def _subtitulo(texto):
        r = fila[0]
        ws.row_dimensions[r].height = 22
        cell = ws.cell(row=r, column=1, value=texto)
        cell.font      = Font(name="Calibri", bold=True, color="375623", size=10)
        cell.fill      = PatternFill("solid", fgColor=COLOR_SECCION)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border    = _border_thin()
        ws.merge_cells(f"A{r}:B{r}")
        fila[0] += 1

    def _item(label, valor, bold=False, bg=None, fmt=None):
        r = fila[0]
        ws.row_dimensions[r].height = 20
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font      = Font(name="Calibri", size=10, bold=bold)
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        c1.border    = _border_thin()
        if bg:
            c1.fill = PatternFill("solid", fgColor=bg)

        c2 = ws.cell(row=r, column=2, value=valor)
        c2.font         = Font(name="Calibri", size=10, bold=bold)
        c2.number_format = fmt if fmt else fmt_num
        c2.alignment    = Alignment(horizontal="right", vertical="center")
        c2.border       = _border_thin()
        if bg:
            c2.fill = PatternFill("solid", fgColor=bg)
        fila[0] += 1

    def _espacio():
        fila[0] += 1

    # ── Cálculos ──────────────────────────────────────────────────────────
    alicuota_decimal     = alicuota_pct / 100.0
    impuesto_determinado = ingresos_declarados * alicuota_decimal
    saldo_dj             = impuesto_determinado - total_deducciones
    # El saldo anterior actúa como crédito: reduce la deuda (o acumula el favor)
    saldo_acumulado      = saldo_dj - saldo_anterior_iibb

    a_pagar = saldo_acumulado if saldo_acumulado > 0 else 0.0
    a_favor = abs(saldo_acumulado) if saldo_acumulado <= 0 else 0.0

    # ── Construcción de la hoja ───────────────────────────────────────────
    _titulo("POSICIÓN DE INGRESOS BRUTOS  ·  R-606M  ·  Impuesto sobre los Ingresos Brutos — ARBA")

    _subtitulo("Ingresos declarados / Impuesto determinado")
    _item("Ingresos declarados  (Neto 21% + Neto 10,5%  del Libro IVA Ventas)",
          ingresos_declarados)
    _item("Alícuota IIBB (%)",
          alicuota_pct,
          fmt=fmt_pct)
    _item("Total Impuesto determinado / mínimo",
          impuesto_determinado,
          bold=True,
          bg=COLOR_TOTAL)

    _espacio()

    _subtitulo("Deducciones computadas para el período")
    _item("Total de deducciones  (RET. ING. BRUTOS del Libro IVA Compras)",
          total_deducciones,
          bold=True,
          bg=COLOR_TOTAL)

    _espacio()

    _subtitulo("Posición del período")
    _item("Saldo de la DJ  (Impuesto determinado − Total deducciones)",
          saldo_dj)
    _item("Saldo a favor anterior IIBB",
          saldo_anterior_iibb)
    _item("Saldo acumulado al cierre  (Saldo de la DJ + Saldo a favor anterior)",
          saldo_acumulado,
          bold=True,
          bg=COLOR_TOTAL)

    _espacio()

    _subtitulo("Resultado")
    if a_pagar > 0:
        _item("A PAGAR IIBB",
              a_pagar,
              bold=True,
              bg=COLOR_PAGAR)
        _item("IIBB A FAVOR",
              0.0)
    else:
        _item("A PAGAR IIBB",
              0.0)
        _item("IIBB A FAVOR",
              a_favor,
              bold=True,
              bg=COLOR_FAVOR)


def _verificar_ret_iibb(
    compras_con_ret: list[dict],
    deducciones_arba: dict,
    log_fn=None,
) -> None:
    """
    Corrobora que la columna RET. ING. BRUTOS del libro IVA COMPRAS
    coincida con el detalle ARBA: PERCEPCION + SIRTAC + BANCARIAS/PSP.
    Emite un log con el resultado; lanza una advertencia si hay diferencia.
    """
    def _log(m):
        if log_fn:
            log_fn(m)

    # Total en el libro generado
    total_libro = sum(r.get("ret_iibb", 0) or 0 for r in compras_con_ret)

    # Totales ARBA por componente
    percepciones = deducciones_arba.get("percepciones", [])
    sitrac       = deducciones_arba.get("sitrac", [])
    bancarias    = deducciones_arba.get("retenciones_bancarias", [])

    total_percepcion = sum(x.get("importe", 0) or 0 for x in percepciones)
    total_sirtac     = sum(x.get("importe", 0) or 0 for x in sitrac)
    total_bancarias  = sum(x.get("importe", 0) or 0 for x in bancarias)
    total_arba       = total_percepcion + total_sirtac + total_bancarias

    diferencia = abs(total_libro - total_arba)
    ok = diferencia < 0.02  # tolerancia de 2 centavos por redondeo

    _log("")
    _log("═" * 60)
    _log("  VERIFICACIÓN RET. ING. BRUTOS (IVA COMPRAS)")
    _log("═" * 60)
    _log(f"  ARBA - Percepción IIBB :  $ {total_percepcion:>16,.2f}")
    _log(f"  ARBA - SIRTAC          :  $ {total_sirtac:>16,.2f}")
    _log(f"  ARBA - Bancarias/PSP   :  $ {total_bancarias:>16,.2f}")
    _log(f"  {'─'*40}")
    _log(f"  TOTAL ARBA             :  $ {total_arba:>16,.2f}")
    _log(f"  TOTAL libro IVA COMPRAS:  $ {total_libro:>16,.2f}")
    _log(f"  {'─'*40}")
    if ok:
        _log(f"  ✅ OK — Diferencia: $0.00  (cuadra exacto)")
    else:
        _log(f"  ❌ DIFERENCIA: $ {diferencia:>12,.2f}  ← REVISAR")
    _log("═" * 60)
